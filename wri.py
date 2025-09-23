import os
import re
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import json
import requests
import time  # ✅ Added for retries

# === URL ===
URL = "https://wri-india.org/about/procurement-opportunities"

# ✅ Enhanced headers for better success in cloud
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": "https://www.google.com/"
}

# === Load keywords from keywords.json ===
def load_keywords_from_json(filename):
    """Loads the keywords from the JSON file."""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"⚠️ Warning: '{filename}' not found. No verticals will be matched.")
        return {}

# === Custom keywords for How_to_Apply ===
custom_keywords = [
    "Selection Criteria", "Evaluation & Follow-Up", "Application Guidelines", "Eligible Applicants:",
    "Scope of Work:", "Proposal Requirements", "Evaluation Criteria", "Submission Details", "Eligible Entities",
    "How to apply", "Purpose of RFP", "Proposal Guidelines", "Eligibility Criteria", "Application must include:",
    "Eligibility", "Submission of Tender:", "Technical Bid-", "Who Can Apply", "Documents Required", "Expectation:",
    "Eligibility Criterion:", "Submission terms:", "Vendor Qualifications", "To apply",
    "To know about the eligibility criteria:", "The agency's specific responsibilities include –",
    "SELCO Foundation will be responsible for:", "Partner Eligibility Criteria", "Proposal Submission Requirements",
    "Proposal Evaluation Criteria", "Eligibility Criteria for CSOs to be part of the programme:", "Pre-Bid Queries:",
    "Response to Pre-Bid Queries:", "Submission of Bid:", "Applicant Profiles:", "What we like to see in grant applications:",
    "Research that is supported by the SVRI must:", "Successful projects are most often:", "Criteria for funding:",
    "Before you begin to write your proposal, consider that IEF prefers to fund:",
    "As you prepare your budget, these are some items that IEF will not fund:", "Organizational Profile",
    "Selection Process", "Proposal Submission Guidelines", "Terms and Conditions", "Security Deposit:",
    "Facilities and Support Offered under the call for proposal:", "Prospective Consultants should demonstrate:"
]

def extract_how_to_apply(description: str) -> str:
    """Find heading + following paragraph(s) for matching keywords."""
    if not description or not isinstance(description, str):
        return "N/A"

    norm_keywords = [kw.lower().rstrip(":") for kw in custom_keywords]

    # Split description into segments (sentences/paragraphs)
    segments = re.split(r'(\.\s+|\n+)', description)
    segments = [s.strip() for s in segments if s.strip() and not s.strip().startswith('.')]

    matched_sections = []
    i = 0
    while i < len(segments):
        segment = segments[i]
        seg_lower = segment.lower()

        if any(kw in seg_lower for kw in norm_keywords):
            section = ["• " + segment]
            i += 1
            while i < len(segments):
                next_segment = segments[i]
                next_lower = next_segment.lower()
                if any(kw in next_lower for kw in norm_keywords):
                    break
                section.append(next_segment)
                i += 1
            matched_sections.append(" ".join(section))
        else:
            i += 1

    return "\n".join(matched_sections).strip() or "N/A"

def find_matched_vertical(title: str, description: str, keywords_data: dict) -> str:
    matched_verticals = []
    text_to_check = f"{title.lower()} {description.lower()}"
    
    for vertical, keywords in keywords_data.items():
        if any(keyword.lower() in text_to_check for keyword in keywords):
            matched_verticals.append(vertical)
            
    return ", ".join(matched_verticals) if matched_verticals else "N/A"

def fetch_wri_opportunities():
    listings = []
    
    keywords_data = load_keywords_from_json("keywords.json")

    response = None
    for attempt in range(3):  # ✅ Add retries
        try:
            response = requests.get(URL, headers=HEADERS, timeout=30, verify=False)  # ✅ Increased timeout
            print(f"✅ WRI response status: {response.status_code} (Attempt {attempt+1})")
            response.raise_for_status()
            break
        except requests.exceptions.RequestException as e:
            print(f"⚠️ WRI fetch attempt {attempt+1} failed: {e}")
            if attempt < 2:
                time.sleep(2)
            else:
                print("❌ Max retries reached for WRI. Returning empty list.")
                return listings

    try:
        soup = BeautifulSoup(response.text, "html.parser")
    except Exception as e:
        print(f"❌ Failed to parse WRI page: {e}")
        return listings

    titles = soup.find_all("div", class_="field--name-field-title")
    descriptions = soup.find_all("div", class_="field--name-field-body")

    for t, d in zip(titles, descriptions):
        title = t.get_text(strip=True)
        description = d.get_text(separator=" ", strip=True)

        # --- improved link extraction ---
        link = URL
        anchors = d.find_all("a", href=True)

        if anchors:
            hrefs = [a["href"].strip() for a in anchors if a["href"].strip()]
            pdfs = [h for h in hrefs if h.lower().endswith(".pdf")]
            if pdfs:
                link = pdfs[0]
            else:
                http_links = [h for h in hrefs if h.startswith("http")]
                if http_links:
                    link = http_links[0]
                else:
                    rel_links = [h for h in hrefs if h.startswith("/")]
                    if rel_links:
                        link = "https://wri-india.org" + rel_links[0]

        if link.startswith("//"):
            link = "https:" + link

        how_to_apply = extract_how_to_apply(description)
        matched_vertical = find_matched_vertical(title, description, keywords_data)

        listings.append({
            "Source": "WRI",
            "Type": pd.NA,
            "Title": title,
            "Description": description,
            "How_to_Apply": how_to_apply,
            "Matched_Vertical": matched_vertical,
            "Deadline": pd.NaT,
            "Days_Left": pd.NA,
            "Clickable_Link": '=HYPERLINK("{}","{}")'.format(link.replace('"', '""'), title.replace('"', '""'))
        })

    print(f"✅ WRI scraped {len(listings)} items")
    return listings

def run_scraper():
    data = fetch_wri_opportunities()
    if not data:
        print("⚠️ No data found.")
        return

    output_dir = os.path.join(os.getcwd(), "output")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    df = pd.DataFrame(data)

    # ✅ Guarantee schema
    final_columns = [
        "Source", "Type", "Title", "Description",
        "How_to_Apply", "Matched_Vertical", "Deadline",
        "Days_Left", "Clickable_Link"
    ]
    df = df.reindex(columns=final_columns)

    excel_path = os.path.join(output_dir, "wri_opportunities.xlsx")
    df.to_excel(excel_path, index=False, engine="openpyxl")

    wb = load_workbook(excel_path)
    ws = wb.active
    for col, width in {"A": 15, "B": 15, "C": 50, "D": 80, "E": 80, "F": 25, "G": 18, "H": 12, "I": 50}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(excel_path)

    print(f"✅ Excel saved to {excel_path}")

if __name__ == "__main__":
    run_scraper()