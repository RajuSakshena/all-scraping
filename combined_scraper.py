import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# Import all scrapers
from main_scraper import scrape_ngobox
from dev import scrape_devnetjobs
from nasscom import scrape_nasscom
from wri import fetch_wri_opportunities
from hcl import scrape_hcl
from metro import fetch_metro_tenders
from niua_tenders import scrape_niua_tenders


def run_combined_scraper(return_df=False):  # ✅ return_df added for API use
    # ---------------- SCRAPERS ----------------
    print("🔍 Running NGOBOX scraper...")
    ngobox_df = scrape_ngobox()

    print("🔍 Running DevNetJobs India scraper...")
    devnet_df = scrape_devnetjobs()

    print("🔍 Running Nasscom scraper...")
    nasscom_df = scrape_nasscom()

    print("🔍 Running WRI scraper...")
    try:
        wri_data = fetch_wri_opportunities()
        if wri_data:
            wri_df = pd.DataFrame(wri_data)
            wri_df["Source"] = "WRI"
            wri_df["Type"] = "N/A"
            wri_df["Deadline"] = pd.NaT
            wri_df["Days_Left"] = pd.NA
            print(f"  -> WRI scraped {len(wri_df)} items.")
        else:
            wri_df = pd.DataFrame()
            print("⚠️ WRI returned no data.")
    except Exception as e:
        print(f"❌ WRI scraper failed: {e}")
        wri_df = pd.DataFrame()

    print("🔍 Running HCL Foundation scraper...")
    try:
        hcl_df = scrape_hcl()
        print(f"  -> HCL scraped {len(hcl_df)} items." if not hcl_df.empty else "⚠️ HCL returned no data.")
    except Exception as e:
        print(f"❌ HCL scraper failed: {e}")
        hcl_df = pd.DataFrame()

    print("🔍 Running Nagpur Metro Rail scraper...")
    try:
        metro_df = fetch_metro_tenders()
        print(f"  -> Metro scraped {len(metro_df)} items." if not metro_df.empty else "⚠️ Metro returned no data.")
    except Exception as e:
        print(f"❌ Metro scraper failed: {e}")
        metro_df = pd.DataFrame()

    print("🔍 Running NIUA Tenders scraper...")
    try:
        niua_raw_df = scrape_niua_tenders()
        if niua_raw_df.empty:
            niua_df = pd.DataFrame()
            print("⚠️ NIUA returned no data.")
        else:
            niua_df = pd.DataFrame({
                "Source": "NIUA",
                "Type": "Tender",
                "Title": niua_raw_df["Tender_Title"],
                "Description": pd.NA,
                "How_to_Apply": pd.NA,
                "Matched_Vertical": pd.NA,
                "Deadline": niua_raw_df["Submission_Deadline"],
                "Days_Left": pd.NA,
                "Clickable_Link": niua_raw_df["Tender_Link"]
            })
            print(f"  -> NIUA scraped {len(niua_df)} items.")
    except Exception as e:
        print(f"❌ NIUA scraper failed: {e}")
        niua_df = pd.DataFrame()

    # ---------------- SCHEMA ----------------
    final_columns = [
        "Source", "Type", "Title", "Description",
        "How_to_Apply", "Matched_Vertical", "Deadline",
        "Days_Left", "Clickable_Link"
    ]

    ngobox_df = ngobox_df.reindex(columns=final_columns)
    devnet_df = devnet_df.reindex(columns=final_columns)
    nasscom_df = nasscom_df.reindex(columns=final_columns)
    wri_df = wri_df.reindex(columns=final_columns)
    hcl_df = hcl_df.reindex(columns=final_columns)
    metro_df = metro_df.reindex(columns=final_columns)
    niua_df = niua_df.reindex(columns=final_columns)

    combined_df = pd.concat(
        [ngobox_df, devnet_df, nasscom_df, wri_df, hcl_df, metro_df, niua_df],
        ignore_index=True
    )

    if combined_df.empty:
        print("❌ No data found from any source.")
        return None

    # ---------------- CLEANING ----------------
    def clean_clickable_link(link):
        if pd.isna(link):
            return ""
        link_str = str(link)
        if link_str.upper().startswith("=HYPERLINK"):
            try:
                return link_str.split('"')[1]
            except:
                return link_str
        return link_str

    combined_df["Clickable_Link"] = combined_df["Clickable_Link"].apply(clean_clickable_link)

    # ---------------- 🔥 KEY FIX ----------------
    # ✅ Keep FULL description
    combined_df["Full_Description"] = combined_df["Description"]

    # ✅ Excel-only truncate
    def truncate_description(desc):
        if pd.isna(desc):
            return ""
        desc = str(desc)
        return desc[:300].rstrip() + " ... Read More" if len(desc) > 300 else desc

    combined_df["Description"] = combined_df["Description"].apply(truncate_description)

    # ---------------- FILTER ----------------
    combined_df["Days_Left"] = pd.to_numeric(combined_df["Days_Left"], errors="coerce")
    combined_df = combined_df[combined_df["Days_Left"].fillna(999) >= 0]
    combined_df = combined_df.sort_values(["Days_Left"], ascending=True, na_position="last")

    # ---------------- SAVE EXCEL ----------------
    excel_df = combined_df.drop(columns=["Full_Description"], errors="ignore")

    excel_path = "all_grants.xlsx"
    excel_df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for col, width in zip("ABCDEFGHI", [15,15,50,100,60,25,18,12,60]):
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(excel_path)

    print(f"✅ Excel saved: {excel_path} ({len(excel_df)} rows)")

    # ---------------- RETURN FOR API ----------------
    if return_df:
        api_df = combined_df.copy()
        api_df["Description"] = api_df["Full_Description"]
        api_df = api_df.drop(columns=["Full_Description"], errors="ignore")
        return api_df

    return None


if __name__ == "__main__":
    run_combined_scraper()
